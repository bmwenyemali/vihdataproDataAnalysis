"""
Excel Data Analysis Enhancement Script - Version 2
Author: Bienvenu Mwenyemali
Description: Creates Excel workbook with indicator dropdown and dynamic analysis
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

print("Loading data...")
df = pd.read_excel('datavih.xlsx')
print(f"Data loaded: {df.shape[0]} rows x {df.shape[1]} columns")

# Create new workbook
wb = Workbook()

# Define styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
title_font = Font(bold=True, size=14, color="2E75B6")
kpi_font = Font(bold=True, size=16, color="2E75B6")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')

def add_data_with_style(ws, df, start_row=1, start_col=1):
    """Add dataframe to worksheet with styling"""
    for c_idx, col_name in enumerate(df.columns, start_col):
        cell = ws.cell(row=start_row, column=c_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    for r_idx, row in enumerate(df.values, start_row + 1):
        for c_idx, value in enumerate(row, start_col):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border

# ============================================================
# SHEET 1: Liste des Indicateurs (Reference)
# ============================================================
print("Creating Indicators Reference sheet...")
ws_ref = wb.active
ws_ref.title = "Liste_Indicateurs"

ws_ref['A1'] = "LISTE DES INDICATEURS VIH/SIDA"
ws_ref['A1'].font = title_font
ws_ref.merge_cells('A1:C1')

# Get unique indicators, removing NaN values
indicators = sorted([i for i in df['indicateurs'].unique() if pd.notna(i)])
ws_ref['A3'] = "N°"
ws_ref['B3'] = "Indicateur"
ws_ref['C3'] = "Catégorie"
ws_ref['A3'].font = header_font
ws_ref['A3'].fill = header_fill
ws_ref['B3'].font = header_font
ws_ref['B3'].fill = header_fill
ws_ref['C3'].font = header_font
ws_ref['C3'].fill = header_fill

for idx, ind in enumerate(indicators, 1):
    ws_ref.cell(row=idx+3, column=1, value=idx).border = border
    ws_ref.cell(row=idx+3, column=2, value=ind).border = border
    # Categorize
    if 'test' in ind.lower():
        cat = "Dépistage"
    elif 'tar' in ind.lower() or 'traitement' in ind.lower():
        cat = "Traitement"
    elif 'charge virale' in ind.lower():
        cat = "Suppression Virale"
    elif 'préservatif' in ind.lower():
        cat = "Prévention"
    elif 'vih+' in ind.lower() or 'diagnostiq' in ind.lower():
        cat = "Diagnostic"
    else:
        cat = "Autre"
    ws_ref.cell(row=idx+3, column=3, value=cat).border = border

ws_ref.column_dimensions['A'].width = 5
ws_ref.column_dimensions['B'].width = 70
ws_ref.column_dimensions['C'].width = 20

# Define named range for indicators
indicator_range = f"Liste_Indicateurs!$B$4:$B${3+len(indicators)}"

# ============================================================
# SHEET 2: Cascade ONUSIDA 95-95-95 (Keep this - it's correct)
# ============================================================
print("Creating UNAIDS 95-95-95 Cascade sheet...")
ws_cascade = wb.create_sheet("Cascade_95-95-95")

ws_cascade['A1'] = "CASCADE ONUSIDA 95-95-95 - RDC"
ws_cascade['A1'].font = Font(bold=True, size=16, color="2E75B6")
ws_cascade.merge_cells('A1:H1')

ws_cascade['A2'] = "Objectif: 95% diagnostiqués | 95% sous traitement | 95% charge virale supprimée"
ws_cascade['A2'].font = Font(italic=True, size=10)
ws_cascade.merge_cells('A2:H2')

# Cascade indicators
cascade_indicators = {
    'Testés': 'Nombre de clients testés',
    'Diagnostiqués VIH+': 'Nombre de clients diagnostiqués VIH+',
    'Sous TAR': 'Nombre de PVVIH sous TAR',
    'Charge Virale Supprimée': 'Nombre  de PVVIH sous TAR qui ont supprimée la charge virale'
}

# By Year
ws_cascade['A4'] = "Cascade par Année"
ws_cascade['A4'].font = Font(bold=True, size=12)

cascade_year = []
for year in sorted(df['annees'].unique()):
    year_data = df[df['annees'] == year]
    row = {'Année': int(year)}
    for name, indicator in cascade_indicators.items():
        row[name] = int(year_data[year_data['indicateurs'] == indicator]['Valeur'].sum())
    cascade_year.append(row)

cascade_year_df = pd.DataFrame(cascade_year)
add_data_with_style(ws_cascade, cascade_year_df, start_row=5)

# Add cascade rates
rate_row = 5 + len(cascade_year_df) + 2
ws_cascade.cell(row=rate_row, column=1, value="Année").font = header_font
ws_cascade.cell(row=rate_row, column=1).fill = header_fill
ws_cascade.cell(row=rate_row, column=2, value="Taux Traitement (%)").font = header_font
ws_cascade.cell(row=rate_row, column=2).fill = header_fill
ws_cascade.cell(row=rate_row, column=3, value="Taux Suppression (%)").font = header_font
ws_cascade.cell(row=rate_row, column=3).fill = header_fill

for idx, row in enumerate(cascade_year):
    r = rate_row + 1 + idx
    ws_cascade.cell(row=r, column=1, value=row['Année']).border = border
    treat_rate = round(100 * row['Sous TAR'] / row['Diagnostiqués VIH+'], 2) if row['Diagnostiqués VIH+'] > 0 else 0
    supp_rate = round(100 * row['Charge Virale Supprimée'] / row['Sous TAR'], 2) if row['Sous TAR'] > 0 else 0
    ws_cascade.cell(row=r, column=2, value=treat_rate).border = border
    ws_cascade.cell(row=r, column=3, value=supp_rate).border = border

# Cascade chart
chart = BarChart()
chart.type = "col"
chart.title = "Cascade ONUSIDA par Année"
chart.y_axis.title = "Nombre de Personnes"
chart.x_axis.title = "Année"

data = Reference(ws_cascade, min_col=2, min_row=5, max_row=5+len(cascade_year_df), max_col=5)
cats = Reference(ws_cascade, min_col=1, min_row=6, max_row=5+len(cascade_year_df))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width = 18
chart.height = 12
ws_cascade.add_chart(chart, "G4")

# By Province
prov_start = rate_row + len(cascade_year) + 4
ws_cascade.cell(row=prov_start, column=1, value="Cascade par Province").font = Font(bold=True, size=12)

cascade_prov = []
for prov in sorted(df['provinces'].unique()):
    prov_data = df[df['provinces'] == prov]
    row = {'Province': prov}
    for name, indicator in cascade_indicators.items():
        row[name] = int(prov_data[prov_data['indicateurs'] == indicator]['Valeur'].sum())
    cascade_prov.append(row)

cascade_prov_df = pd.DataFrame(cascade_prov)
cascade_prov_df = cascade_prov_df.sort_values('Testés', ascending=False)
add_data_with_style(ws_cascade, cascade_prov_df, start_row=prov_start+1)

# Column widths
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_cascade.column_dimensions[col].width = 22

# ============================================================
# SHEETS 3-12: Analysis per Indicator (10 most important)
# ============================================================
print("Creating analysis sheets for top indicators...")

# Select top 10 most important indicators
important_indicators = [
    'Nombre de clients testés',
    'Nombre de clients diagnostiqués VIH+',
    'Nombre de PVVIH sous TAR',
    'Nombre  de PVVIH sous TAR qui ont supprimée la charge virale',
    'Nombre de préservatifs masculins distribués',
    'Nombre de femmes enceintes séropositives ayant reçu des ARV pour la PTME',
    'Nombre de nouveau-nés exposés ayant bénéficié de la PCR dans les deux mois de vie',
    'Nombre de clients testés (Clinique)',
    'Nombre de clients testés (Communautaire)',
    'Nombre de préservatifs féminins distribués'
]

# Filter to indicators that exist in data
available_indicators = [ind for ind in important_indicators if ind in df['indicateurs'].values]

for ind_idx, indicator in enumerate(available_indicators[:10], 1):
    print(f"  Creating sheet for: {indicator[:50]}...")
    
    # Create safe sheet name (max 31 chars)
    sheet_name = f"Ind_{ind_idx}"
    ws = wb.create_sheet(sheet_name)
    
    # Filter data for this indicator
    ind_df = df[df['indicateurs'] == indicator].copy()
    
    # Title
    ws['A1'] = f"ANALYSE: {indicator}"
    ws['A1'].font = Font(bold=True, size=12, color="2E75B6")
    ws.merge_cells('A1:F1')
    
    # Total KPI
    total_value = ind_df['Valeur'].sum()
    ws['A3'] = "TOTAL"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = int(total_value)
    ws['B3'].font = kpi_font
    ws['B3'].number_format = '#,##0'
    
    # ---- Analysis by Year ----
    ws['A5'] = "Analyse par Année"
    ws['A5'].font = Font(bold=True, size=11, color="2E75B6")
    
    year_agg = ind_df.groupby('annees')['Valeur'].sum().reset_index()
    year_agg.columns = ['Année', 'Valeur']
    year_agg['Année'] = year_agg['Année'].astype(int)
    year_agg['Valeur'] = year_agg['Valeur'].astype(int)
    
    # YoY growth
    year_agg['Croissance (%)'] = year_agg['Valeur'].pct_change() * 100
    year_agg['Croissance (%)'] = year_agg['Croissance (%)'].round(2).fillna(0)
    
    add_data_with_style(ws, year_agg, start_row=6)
    
    # Year chart
    chart = LineChart()
    chart.title = "Évolution Annuelle"
    chart.y_axis.title = "Valeur"
    data = Reference(ws, min_col=2, min_row=6, max_row=6+len(year_agg))
    cats = Reference(ws, min_col=1, min_row=7, max_row=6+len(year_agg))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 12
    chart.height = 8
    ws.add_chart(chart, "E5")
    
    # ---- Analysis by Province ----
    prov_row = 6 + len(year_agg) + 3
    ws.cell(row=prov_row, column=1, value="Analyse par Province").font = Font(bold=True, size=11, color="2E75B6")
    
    prov_agg = ind_df.groupby('provinces')['Valeur'].sum().reset_index()
    prov_agg.columns = ['Province', 'Valeur']
    prov_agg = prov_agg.sort_values('Valeur', ascending=False)
    prov_agg['Valeur'] = prov_agg['Valeur'].astype(int)
    prov_agg['Rang'] = range(1, len(prov_agg) + 1)
    prov_agg['% du Total'] = (prov_agg['Valeur'] / prov_agg['Valeur'].sum() * 100).round(2)
    prov_agg = prov_agg[['Rang', 'Province', 'Valeur', '% du Total']]
    
    add_data_with_style(ws, prov_agg, start_row=prov_row+1)
    
    # Province bar chart (top 10)
    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "Top 10 Provinces"
    data = Reference(ws, min_col=3, min_row=prov_row+1, max_row=min(prov_row+11, prov_row+1+len(prov_agg)))
    cats = Reference(ws, min_col=2, min_row=prov_row+2, max_row=min(prov_row+11, prov_row+1+len(prov_agg)))
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.width = 14
    chart2.height = 10
    ws.add_chart(chart2, "F" + str(prov_row))
    
    # ---- Analysis by Quarter ----
    trim_row = prov_row + len(prov_agg) + 4
    ws.cell(row=trim_row, column=1, value="Analyse par Trimestre").font = Font(bold=True, size=11, color="2E75B6")
    
    trim_agg = ind_df.groupby('trimestres')['Valeur'].sum().reset_index()
    trim_agg.columns = ['Trimestre', 'Valeur']
    trim_agg['Valeur'] = trim_agg['Valeur'].astype(int)
    trim_agg['% du Total'] = (trim_agg['Valeur'] / trim_agg['Valeur'].sum() * 100).round(2)
    
    add_data_with_style(ws, trim_agg, start_row=trim_row+1)
    
    # ---- Analysis by Gender ----
    gender_row = trim_row + len(trim_agg) + 4
    ws.cell(row=gender_row, column=1, value="Analyse par Sexe").font = Font(bold=True, size=11, color="2E75B6")
    
    gender_agg = ind_df.groupby('sexes')['Valeur'].sum().reset_index()
    gender_agg.columns = ['Sexe', 'Valeur']
    gender_agg['Sexe'] = gender_agg['Sexe'].fillna('Non spécifié')
    gender_agg['Valeur'] = gender_agg['Valeur'].astype(int)
    gender_agg['% du Total'] = (gender_agg['Valeur'] / gender_agg['Valeur'].sum() * 100).round(2)
    
    add_data_with_style(ws, gender_agg, start_row=gender_row+1)
    
    # Gender pie chart
    if len(gender_agg) > 1:
        pie = PieChart()
        pie.title = "Répartition par Sexe"
        data = Reference(ws, min_col=2, min_row=gender_row+1, max_row=gender_row+1+len(gender_agg))
        labels = Reference(ws, min_col=1, min_row=gender_row+2, max_row=gender_row+1+len(gender_agg))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.width = 10
        pie.height = 8
        ws.add_chart(pie, "E" + str(gender_row))
    
    # ---- Analysis by Age Group ----
    age_row = gender_row + len(gender_agg) + 4
    ws.cell(row=age_row, column=1, value="Analyse par Tranche d'Âge").font = Font(bold=True, size=11, color="2E75B6")
    
    age_agg = ind_df.groupby('tranches_ages')['Valeur'].sum().reset_index()
    age_agg.columns = ['Tranche d\'Âge', 'Valeur']
    age_agg['Tranche d\'Âge'] = age_agg['Tranche d\'Âge'].fillna('Non spécifié')
    age_agg['Valeur'] = age_agg['Valeur'].astype(int)
    age_agg['% du Total'] = (age_agg['Valeur'] / age_agg['Valeur'].sum() * 100).round(2)
    
    add_data_with_style(ws, age_agg, start_row=age_row+1)
    
    # Age bar chart
    if len(age_agg) > 1:
        chart3 = BarChart()
        chart3.type = "col"
        chart3.title = "Distribution par Âge"
        data = Reference(ws, min_col=2, min_row=age_row+1, max_row=age_row+1+len(age_agg))
        cats = Reference(ws, min_col=1, min_row=age_row+2, max_row=age_row+1+len(age_agg))
        chart3.add_data(data, titles_from_data=True)
        chart3.set_categories(cats)
        chart3.width = 12
        chart3.height = 8
        ws.add_chart(chart3, "E" + str(age_row))
    
    # ---- Cross-tab: Year x Province (Pivot) ----
    pivot_row = age_row + len(age_agg) + 4
    ws.cell(row=pivot_row, column=1, value="Tableau Croisé: Province × Année").font = Font(bold=True, size=11, color="2E75B6")
    
    pivot = ind_df.pivot_table(values='Valeur', index='provinces', columns='annees', aggfunc='sum', fill_value=0)
    pivot['TOTAL'] = pivot.sum(axis=1)
    pivot = pivot.reset_index()
    pivot = pivot.sort_values('TOTAL', ascending=False)
    
    # Convert columns to proper types
    pivot.columns = [str(int(c)) if isinstance(c, float) else str(c) for c in pivot.columns]
    
    add_data_with_style(ws, pivot, start_row=pivot_row+1)
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12

# ============================================================
# SHEET: Dashboard with Dropdown
# ============================================================
print("Creating Interactive Dashboard sheet...")
ws_dash = wb.create_sheet("Dashboard_Interactif", 1)

ws_dash['A1'] = "TABLEAU DE BORD INTERACTIF - VIH/SIDA RDC"
ws_dash['A1'].font = Font(bold=True, size=16, color="2E75B6")
ws_dash.merge_cells('A1:H1')

ws_dash['A2'] = "Sélectionnez un indicateur dans la liste ci-dessous pour voir les analyses correspondantes"
ws_dash['A2'].font = Font(italic=True, size=10)
ws_dash.merge_cells('A2:H2')

# Dropdown cell
ws_dash['A4'] = "SÉLECTIONNER UN INDICATEUR:"
ws_dash['A4'].font = Font(bold=True, size=11)
ws_dash['B4'] = indicators[0]  # Default first indicator
ws_dash['B4'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ws_dash['B4'].font = Font(bold=True)

# Add data validation dropdown
dv = DataValidation(
    type="list",
    formula1=f"'{ws_ref.title}'!$B$4:$B${3+len(indicators)}",
    allow_blank=False
)
dv.prompt = "Sélectionnez un indicateur"
dv.promptTitle = "Choix d'indicateur"
ws_dash.add_data_validation(dv)
dv.add(ws_dash['B4'])

ws_dash.merge_cells('B4:G4')

# Instructions
ws_dash['A6'] = "INSTRUCTIONS:"
ws_dash['A6'].font = Font(bold=True, size=12, color="C00000")

instructions = [
    "1. Cliquez sur la cellule jaune ci-dessus (B4)",
    "2. Un menu déroulant apparaîtra avec la liste des indicateurs",
    "3. Sélectionnez l'indicateur que vous souhaitez analyser",
    "4. Les données détaillées sont dans les onglets Ind_1 à Ind_10",
    "",
    "ONGLETS DE CE CLASSEUR:",
    "• Liste_Indicateurs: Liste complète des 118 indicateurs",
    "• Dashboard_Interactif: Cette page (sélection d'indicateur)",
    "• Cascade_95-95-95: Analyse de la cascade ONUSIDA",
    "• Ind_1 à Ind_10: Analyses détaillées par indicateur",
]

for idx, text in enumerate(instructions, 7):
    ws_dash.cell(row=idx, column=1, value=text)

# Index of indicator sheets
ws_dash['A20'] = "INDEX DES ONGLETS D'ANALYSE PAR INDICATEUR:"
ws_dash['A20'].font = Font(bold=True, size=12)

ws_dash['A21'] = "Onglet"
ws_dash['B21'] = "Indicateur"
ws_dash['A21'].font = header_font
ws_dash['A21'].fill = header_fill
ws_dash['B21'].font = header_font
ws_dash['B21'].fill = header_fill

for idx, ind in enumerate(available_indicators[:10], 1):
    ws_dash.cell(row=21+idx, column=1, value=f"Ind_{idx}").border = border
    ws_dash.cell(row=21+idx, column=2, value=ind).border = border

# Quick stats section
ws_dash['A35'] = "STATISTIQUES RAPIDES DU DATASET:"
ws_dash['A35'].font = Font(bold=True, size=12)

quick_stats = [
    ["Métrique", "Valeur"],
    ["Nombre total d'enregistrements", len(df)],
    ["Nombre de provinces", df['provinces'].nunique()],
    ["Nombre d'indicateurs", df['indicateurs'].nunique()],
    ["Période couverte", f"{int(df['annees'].min())} - {int(df['annees'].max())}"],
]

for r_idx, row in enumerate(quick_stats, 36):
    for c_idx, val in enumerate(row, 1):
        cell = ws_dash.cell(row=r_idx, column=c_idx, value=val)
        cell.border = border
        if r_idx == 36:
            cell.font = header_font
            cell.fill = header_fill

# Column widths
ws_dash.column_dimensions['A'].width = 40
ws_dash.column_dimensions['B'].width = 70

# ============================================================
# SHEET: Formulas Reference (Excel formulas to use)
# ============================================================
print("Creating Formulas Reference sheet...")
ws_form = wb.create_sheet("Formules_Reference")

ws_form['A1'] = "FORMULES EXCEL POUR ANALYSE DYNAMIQUE"
ws_form['A1'].font = title_font
ws_form.merge_cells('A1:D1')

ws_form['A3'] = "Ces formules peuvent être utilisées avec le dropdown pour filtrer les données:"
ws_form['A3'].font = Font(italic=True)
ws_form.merge_cells('A3:D3')

formulas = [
    ["Fonction", "Syntaxe", "Description", "Exemple"],
    ["SUMIF", "=SUMIF(plage_critère, critère, plage_somme)", "Somme conditionnelle", "=SUMIF(indicateurs,B4,Valeur)"],
    ["COUNTIF", "=COUNTIF(plage, critère)", "Comptage conditionnel", "=COUNTIF(indicateurs,B4)"],
    ["AVERAGEIF", "=AVERAGEIF(plage_critère, critère, plage_moyenne)", "Moyenne conditionnelle", "=AVERAGEIF(indicateurs,B4,Valeur)"],
    ["SUMIFS", "=SUMIFS(somme, plage1, crit1, plage2, crit2)", "Somme multi-critères", "=SUMIFS(Valeur,indicateurs,B4,annees,2024)"],
    ["FILTER", "=FILTER(données, condition)", "Filtrage dynamique (Excel 365)", "=FILTER(A:H,C:C=B4)"],
    ["UNIQUE", "=UNIQUE(plage)", "Valeurs uniques", "=UNIQUE(provinces)"],
    ["SORT", "=SORT(plage, col, ordre)", "Tri dynamique", "=SORT(FILTER(...),2,-1)"],
    ["XLOOKUP", "=XLOOKUP(valeur, recherche, retour)", "Recherche moderne", "=XLOOKUP(B4,indicateurs,Valeur)"],
]

for r_idx, row in enumerate(formulas, 5):
    for c_idx, val in enumerate(row, 1):
        cell = ws_form.cell(row=r_idx, column=c_idx, value=val)
        cell.border = border
        if r_idx == 5:
            cell.font = header_font
            cell.fill = header_fill

ws_form.column_dimensions['A'].width = 15
ws_form.column_dimensions['B'].width = 50
ws_form.column_dimensions['C'].width = 35
ws_form.column_dimensions['D'].width = 45

# ============================================================
# Save workbook
# ============================================================
print("\nSaving workbook...")
output_file = 'datavih_analysis.xlsx'
wb.save(output_file)
print(f"✓ Workbook saved as: {output_file}")

print("\n" + "="*60)
print("WORKSHEETS CREATED:")
print("="*60)
print("  1. Liste_Indicateurs - Liste des 118 indicateurs")
print("  2. Dashboard_Interactif - Dropdown pour sélection d'indicateur")
print("  3. Cascade_95-95-95 - Analyse ONUSIDA (correcte)")
for idx, ind in enumerate(available_indicators[:10], 4):
    print(f"  {idx}. Ind_{idx-3} - {ind[:50]}...")
print(f"  {4+len(available_indicators[:10])}. Formules_Reference - Guide des formules Excel")

print("\n" + "="*60)
print("ANALYSES PAR INDICATEUR INCLUENT:")
print("="*60)
print("  • Analyse par Année (avec croissance YoY)")
print("  • Analyse par Province (avec classement)")
print("  • Analyse par Trimestre")
print("  • Analyse par Sexe")
print("  • Analyse par Tranche d'Âge")
print("  • Tableau Croisé Province × Année")
print("  • Graphiques: Ligne, Barres, Camembert")

print("\nDone!")
