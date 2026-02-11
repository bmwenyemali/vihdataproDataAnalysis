"""
Excel Data Analysis Enhancement Script
Author: Bienvenu Mwenyemali
Description: Adds multiple worksheets to datavih.xlsx demonstrating Excel data analysis skills
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo

print("Loading data...")
# Load the original data
df = pd.read_excel('datavih.xlsx')
print(f"Data loaded: {df.shape[0]} rows x {df.shape[1]} columns")

# Create a copy of the workbook
wb = load_workbook('datavih.xlsx')

# Define styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')

def style_header(ws, row=1, start_col=1, end_col=None):
    """Apply header styling to a row"""
    if end_col is None:
        end_col = ws.max_column
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

def add_data_with_style(ws, df, start_row=1, start_col=1):
    """Add dataframe to worksheet with styling"""
    # Add headers
    for c_idx, col_name in enumerate(df.columns, start_col):
        cell = ws.cell(row=start_row, column=c_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Add data
    for r_idx, row in enumerate(df.values, start_row + 1):
        for c_idx, value in enumerate(row, start_col):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border

# ============================================================
# SHEET 1: Summary Statistics (Statistiques Résumées)
# ============================================================
print("Creating Summary Statistics sheet...")
ws_summary = wb.create_sheet("Statistiques_Resume", 0)

# Title
ws_summary['A1'] = "RÉSUMÉ STATISTIQUE - DONNÉES VIH/SIDA RDC"
ws_summary['A1'].font = Font(bold=True, size=16, color="2E75B6")
ws_summary.merge_cells('A1:F1')

# Basic stats
ws_summary['A3'] = "Statistiques de Base"
ws_summary['A3'].font = Font(bold=True, size=12)

stats_data = [
    ["Métrique", "Valeur"],
    ["Nombre total d'enregistrements", len(df)],
    ["Nombre de provinces", df['provinces'].nunique()],
    ["Nombre d'années", df['annees'].nunique()],
    ["Années couvertes", f"{df['annees'].min()} - {df['annees'].max()}"],
    ["Nombre d'indicateurs", df['indicateurs'].nunique()],
    ["Valeur totale", df['Valeur'].sum()],
    ["Valeur moyenne", round(df['Valeur'].mean(), 2)],
    ["Valeur médiane", df['Valeur'].median()],
    ["Valeur maximale", df['Valeur'].max()],
    ["Valeur minimale", df['Valeur'].min()],
    ["Écart-type", round(df['Valeur'].std(), 2)],
]

for r_idx, row in enumerate(stats_data, 4):
    for c_idx, val in enumerate(row, 1):
        cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
        cell.border = border
        if r_idx == 4:
            cell.font = header_font
            cell.fill = header_fill

# Column widths
ws_summary.column_dimensions['A'].width = 35
ws_summary.column_dimensions['B'].width = 25

# ============================================================
# SHEET 2: Province Summary (Résumé par Province)
# ============================================================
print("Creating Province Summary sheet...")
ws_province = wb.create_sheet("Resume_Province", 1)

province_summary = df.groupby('provinces').agg({
    'Valeur': ['sum', 'mean', 'min', 'max', 'count']
}).round(2)
province_summary.columns = ['Total', 'Moyenne', 'Min', 'Max', 'Nb_Records']
province_summary = province_summary.reset_index()
province_summary = province_summary.sort_values('Total', ascending=False)

# Add rank column
province_summary['Rang'] = range(1, len(province_summary) + 1)
province_summary = province_summary[['Rang', 'provinces', 'Total', 'Moyenne', 'Min', 'Max', 'Nb_Records']]

ws_province['A1'] = "RÉSUMÉ PAR PROVINCE"
ws_province['A1'].font = Font(bold=True, size=14, color="2E75B6")
ws_province.merge_cells('A1:G1')

add_data_with_style(ws_province, province_summary, start_row=3)

# Add Excel Table
tab_ref = f"A3:G{len(province_summary) + 3}"
table = Table(displayName="TableProvince", ref=tab_ref)
style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
table.tableStyleInfo = style
ws_province.add_table(table)

# Add bar chart
chart = BarChart()
chart.type = "col"
chart.title = "Top 10 Provinces par Valeur Totale"
chart.y_axis.title = "Valeur Totale"
chart.x_axis.title = "Province"

data = Reference(ws_province, min_col=3, min_row=3, max_row=13, max_col=3)
cats = Reference(ws_province, min_col=2, min_row=4, max_row=13)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4
chart.width = 18
chart.height = 10
ws_province.add_chart(chart, "I3")

# Column widths
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    ws_province.column_dimensions[col].width = 15
ws_province.column_dimensions['B'].width = 20

# ============================================================
# SHEET 3: Year Summary (Résumé par Année)
# ============================================================
print("Creating Year Summary sheet...")
ws_year = wb.create_sheet("Resume_Annee", 2)

year_summary = df.groupby('annees').agg({
    'Valeur': ['sum', 'mean', 'count']
}).round(2)
year_summary.columns = ['Total', 'Moyenne', 'Nb_Records']
year_summary = year_summary.reset_index()

# Calculate YoY growth
year_summary['Croissance_YoY'] = year_summary['Total'].pct_change() * 100
year_summary['Croissance_YoY'] = year_summary['Croissance_YoY'].round(2)
year_summary['Croissance_YoY'] = year_summary['Croissance_YoY'].fillna(0)

ws_year['A1'] = "RÉSUMÉ PAR ANNÉE"
ws_year['A1'].font = Font(bold=True, size=14, color="2E75B6")
ws_year.merge_cells('A1:E1')

add_data_with_style(ws_year, year_summary, start_row=3)

# Add line chart for trend
chart = LineChart()
chart.title = "Évolution Annuelle"
chart.y_axis.title = "Valeur Totale"
chart.x_axis.title = "Année"

data = Reference(ws_year, min_col=2, min_row=3, max_row=3 + len(year_summary))
cats = Reference(ws_year, min_col=1, min_row=4, max_row=3 + len(year_summary))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width = 15
chart.height = 10
ws_year.add_chart(chart, "G3")

# Column widths
for col in ['A', 'B', 'C', 'D', 'E']:
    ws_year.column_dimensions[col].width = 18

# ============================================================
# SHEET 4: UNAIDS 95-95-95 Cascade
# ============================================================
print("Creating UNAIDS Cascade sheet...")
ws_cascade = wb.create_sheet("Cascade_ONUSIDA", 3)

# Calculate cascade data
cascade_indicators = {
    'tested': 'Nombre de clients testés',
    'diagnosed': 'Nombre de clients diagnostiqués VIH+',
    'on_tar': 'Nombre de PVVIH sous TAR',
    'viral_suppressed': 'Nombre  de PVVIH sous TAR qui ont supprimée la charge virale'
}

cascade_data = []
for year in sorted(df['annees'].unique()):
    year_data = df[df['annees'] == year]
    row = {'Année': year}
    for key, indicator in cascade_indicators.items():
        row[key] = year_data[year_data['indicateurs'] == indicator]['Valeur'].sum()
    cascade_data.append(row)

cascade_df = pd.DataFrame(cascade_data)
cascade_df.columns = ['Année', 'Testés', 'Diagnostiqués', 'Sous TAR', 'Charge Virale Supprimée']

# Calculate rates
cascade_df['Taux Traitement (%)'] = (cascade_df['Sous TAR'] / cascade_df['Diagnostiqués'] * 100).round(2)
cascade_df['Taux Suppression (%)'] = (cascade_df['Charge Virale Supprimée'] / cascade_df['Sous TAR'] * 100).round(2)

ws_cascade['A1'] = "CASCADE ONUSIDA 95-95-95"
ws_cascade['A1'].font = Font(bold=True, size=14, color="2E75B6")
ws_cascade.merge_cells('A1:G1')

ws_cascade['A2'] = "Objectifs: 95% Testés | 95% Sous Traitement | 95% Charge Virale Supprimée"
ws_cascade['A2'].font = Font(italic=True, size=10)
ws_cascade.merge_cells('A2:G2')

add_data_with_style(ws_cascade, cascade_df, start_row=4)

# Add bar chart for cascade
chart = BarChart()
chart.type = "col"
chart.title = "Cascade ONUSIDA par Année"
chart.y_axis.title = "Nombre de Personnes"

data = Reference(ws_cascade, min_col=2, min_row=4, max_row=4 + len(cascade_df), max_col=5)
cats = Reference(ws_cascade, min_col=1, min_row=5, max_row=4 + len(cascade_df))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width = 18
chart.height = 12
ws_cascade.add_chart(chart, "I4")

# Column widths
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    ws_cascade.column_dimensions[col].width = 20

# ============================================================
# SHEET 5: Pivot Table - Province by Year
# ============================================================
print("Creating Pivot Table sheet...")
ws_pivot = wb.create_sheet("Tableau_Croise", 4)

pivot_df = df.pivot_table(
    values='Valeur',
    index='provinces',
    columns='annees',
    aggfunc='sum',
    fill_value=0
)
pivot_df['TOTAL'] = pivot_df.sum(axis=1)
pivot_df = pivot_df.reset_index()
pivot_df = pivot_df.sort_values('TOTAL', ascending=False)

ws_pivot['A1'] = "TABLEAU CROISÉ DYNAMIQUE - PROVINCE PAR ANNÉE"
ws_pivot['A1'].font = Font(bold=True, size=14, color="2E75B6")
ws_pivot.merge_cells('A1:G1')

add_data_with_style(ws_pivot, pivot_df, start_row=3)

# Add conditional formatting (color scale)
color_scale = ColorScaleRule(
    start_type='min', start_color='FFFFFF',
    mid_type='percentile', mid_value=50, mid_color='FFC7CE',
    end_type='max', end_color='006400'
)
ws_pivot.conditional_formatting.add(f'B4:F{len(pivot_df) + 3}', color_scale)

# Column widths
ws_pivot.column_dimensions['A'].width = 20
for col in ['B', 'C', 'D', 'E', 'F', 'G']:
    ws_pivot.column_dimensions[col].width = 15

# ============================================================
# SHEET 6: Lookup Reference Table
# ============================================================
print("Creating Lookup Reference sheet...")
ws_lookup = wb.create_sheet("Reference_Lookup", 5)

ws_lookup['A1'] = "TABLES DE RÉFÉRENCE POUR LOOKUP"
ws_lookup['A1'].font = Font(bold=True, size=14, color="2E75B6")
ws_lookup.merge_cells('A1:F1')

# Province codes lookup
ws_lookup['A3'] = "Table 1: Codes Provinces"
ws_lookup['A3'].font = Font(bold=True, size=12)

provinces = sorted(df['provinces'].unique())
province_codes = pd.DataFrame({
    'Code': [f'P{str(i+1).zfill(2)}' for i in range(len(provinces))],
    'Province': provinces,
    'Region': ['Ouest' if p in ['Kinshasa', 'Kongo-Central', 'Kwango', 'Kwilu', 'Mai-Ndombe'] 
               else 'Est' if p in ['Nord-Kivu', 'Sud-Kivu', 'Maniema', 'Ituri'] 
               else 'Nord' if p in ['Équateur', 'Mongala', 'Nord-Ubangi', 'Sud-Ubangi', 'Tshuapa', 'Tshopo']
               else 'Sud' if p in ['Haut-Katanga', 'Haut-Lomami', 'Lualaba', 'Tanganyika']
               else 'Centre' for p in provinces]
})

add_data_with_style(ws_lookup, province_codes, start_row=4)

# Indicator categories lookup
ws_lookup['A' + str(len(provinces) + 7)] = "Table 2: Catégories d'Indicateurs"
ws_lookup['A' + str(len(provinces) + 7)].font = Font(bold=True, size=12)

indicators = df['indicateurs'].unique()
indicator_cat = pd.DataFrame({
    'Indicateur': indicators[:15],  # First 15 indicators
    'Catégorie': ['Dépistage' if 'test' in str(i).lower() 
                  else 'Traitement' if 'tar' in str(i).lower() or 'traitement' in str(i).lower()
                  else 'Prévention' if 'préservatif' in str(i).lower() or 'prévention' in str(i).lower()
                  else 'Suppression' if 'charge' in str(i).lower()
                  else 'Autre' for i in indicators[:15]]
})

add_data_with_style(ws_lookup, indicator_cat, start_row=len(provinces) + 8)

# VLOOKUP/INDEX-MATCH examples
example_row = len(provinces) + 25
ws_lookup['A' + str(example_row)] = "Exemples de Formules LOOKUP"
ws_lookup['A' + str(example_row)].font = Font(bold=True, size=12)

formulas = [
    ["Formule", "Description", "Résultat"],
    ["=VLOOKUP(\"P01\",A5:C30,2,FALSE)", "Trouver province par code", "Bas-Uele"],
    ["=INDEX(B5:B30,MATCH(\"Kinshasa\",B5:B30,0))", "INDEX-MATCH pour province", "Kinshasa"],
    ["=SUMIF(B5:B30,\"Kinshasa\",C5:C30)", "Somme conditionnelle", "N/A"],
    ["=COUNTIF(C5:C30,\"Ouest\")", "Compter par région", "5"],
]

for r_idx, row in enumerate(formulas, example_row + 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws_lookup.cell(row=r_idx, column=c_idx, value=val)
        cell.border = border
        if r_idx == example_row + 1:
            cell.font = header_font
            cell.fill = header_fill

# Column widths
ws_lookup.column_dimensions['A'].width = 45
ws_lookup.column_dimensions['B'].width = 35
ws_lookup.column_dimensions['C'].width = 20

# ============================================================
# SHEET 7: Dashboard Summary
# ============================================================
print("Creating Dashboard sheet...")
ws_dashboard = wb.create_sheet("Dashboard", 6)

# Title
ws_dashboard['A1'] = "TABLEAU DE BORD VIH/SIDA - RDC"
ws_dashboard['A1'].font = Font(bold=True, size=18, color="2E75B6")
ws_dashboard.merge_cells('A1:H1')

ws_dashboard['A2'] = f"Période: {df['annees'].min()} - {df['annees'].max()} | Dernière mise à jour: Février 2026"
ws_dashboard['A2'].font = Font(italic=True, size=10)
ws_dashboard.merge_cells('A2:H2')

# KPI Boxes
kpi_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
kpi_font = Font(bold=True, size=20, color="2E75B6")
kpi_title_font = Font(bold=True, size=10)

# KPI 1: Total Records
ws_dashboard['A4'] = "Total Enregistrements"
ws_dashboard['A4'].font = kpi_title_font
ws_dashboard['A5'] = len(df)
ws_dashboard['A5'].font = kpi_font
ws_dashboard['A5'].fill = kpi_fill

# KPI 2: Total Value
ws_dashboard['C4'] = "Valeur Totale"
ws_dashboard['C4'].font = kpi_title_font
ws_dashboard['C5'] = int(df['Valeur'].sum())
ws_dashboard['C5'].font = kpi_font
ws_dashboard['C5'].fill = kpi_fill
ws_dashboard['C5'].number_format = '#,##0'

# KPI 3: Provinces
ws_dashboard['E4'] = "Provinces Couvertes"
ws_dashboard['E4'].font = kpi_title_font
ws_dashboard['E5'] = df['provinces'].nunique()
ws_dashboard['E5'].font = kpi_font
ws_dashboard['E5'].fill = kpi_fill

# KPI 4: Indicators
ws_dashboard['G4'] = "Indicateurs Suivis"
ws_dashboard['G4'].font = kpi_title_font
ws_dashboard['G5'] = df['indicateurs'].nunique()
ws_dashboard['G5'].font = kpi_font
ws_dashboard['G5'].fill = kpi_fill

# Top 5 Provinces mini-table
ws_dashboard['A8'] = "Top 5 Provinces"
ws_dashboard['A8'].font = Font(bold=True, size=12)

top5 = df.groupby('provinces')['Valeur'].sum().nlargest(5).reset_index()
top5.columns = ['Province', 'Valeur']
top5['Rang'] = range(1, 6)
top5 = top5[['Rang', 'Province', 'Valeur']]

add_data_with_style(ws_dashboard, top5, start_row=9)

# Year totals mini-table
ws_dashboard['E8'] = "Totaux par Année"
ws_dashboard['E8'].font = Font(bold=True, size=12)

year_totals = df.groupby('annees')['Valeur'].sum().reset_index()
year_totals.columns = ['Année', 'Total']

add_data_with_style(ws_dashboard, year_totals, start_row=9, start_col=5)

# Add pie chart for top provinces
pie_data = df.groupby('provinces')['Valeur'].sum().nlargest(5)
pie_df = pie_data.reset_index()

# Add data for pie chart
ws_dashboard['A17'] = "Données pour Graphique"
ws_dashboard['A17'].font = Font(bold=True, size=10)
for idx, (province, value) in enumerate(zip(pie_df['provinces'], pie_df['Valeur']), 18):
    ws_dashboard.cell(row=idx, column=1, value=province)
    ws_dashboard.cell(row=idx, column=2, value=value)

# Create pie chart
pie = PieChart()
pie.title = "Répartition Top 5 Provinces"
labels = Reference(ws_dashboard, min_col=1, min_row=18, max_row=22)
data = Reference(ws_dashboard, min_col=2, min_row=17, max_row=22)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.width = 12
pie.height = 10
ws_dashboard.add_chart(pie, "A24")

# Line chart for yearly trend
line = LineChart()
line.title = "Tendance Annuelle"
line.y_axis.title = "Valeur"
line.x_axis.title = "Année"

data = Reference(ws_dashboard, min_col=6, min_row=9, max_row=9 + len(year_totals))
cats = Reference(ws_dashboard, min_col=5, min_row=10, max_row=9 + len(year_totals))
line.add_data(data, titles_from_data=True)
line.set_categories(cats)
line.width = 12
line.height = 10
ws_dashboard.add_chart(line, "E24")

# Column widths
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_dashboard.column_dimensions[col].width = 15

# ============================================================
# SHEET 8: Data Analysis Formulas
# ============================================================
print("Creating Formulas Reference sheet...")
ws_formulas = wb.create_sheet("Formules_Analyse", 7)

ws_formulas['A1'] = "RÉFÉRENCE DES FORMULES D'ANALYSE EXCEL"
ws_formulas['A1'].font = Font(bold=True, size=14, color="2E75B6")
ws_formulas.merge_cells('A1:D1')

formulas_list = [
    ["Catégorie", "Formule", "Description", "Exemple"],
    ["Statistiques", "=AVERAGE(range)", "Moyenne", "=AVERAGE(Valeur)"],
    ["Statistiques", "=MEDIAN(range)", "Médiane", "=MEDIAN(Valeur)"],
    ["Statistiques", "=STDEV(range)", "Écart-type", "=STDEV(Valeur)"],
    ["Statistiques", "=VAR(range)", "Variance", "=VAR(Valeur)"],
    ["Statistiques", "=PERCENTILE(range, k)", "Percentile", "=PERCENTILE(Valeur, 0.75)"],
    ["Agrégation", "=SUMIF(range, criteria, sum_range)", "Somme conditionnelle", "=SUMIF(provinces,\"Kinshasa\",Valeur)"],
    ["Agrégation", "=SUMIFS(sum, range1, crit1, range2, crit2)", "Somme multi-critères", "=SUMIFS(Valeur,provinces,\"Kinshasa\",annees,2024)"],
    ["Agrégation", "=COUNTIF(range, criteria)", "Comptage conditionnel", "=COUNTIF(provinces,\"Kinshasa\")"],
    ["Agrégation", "=AVERAGEIF(range, criteria, avg_range)", "Moyenne conditionnelle", "=AVERAGEIF(provinces,\"Kinshasa\",Valeur)"],
    ["Lookup", "=VLOOKUP(value, table, col, exact)", "Recherche verticale", "=VLOOKUP(\"P01\",A:C,2,FALSE)"],
    ["Lookup", "=HLOOKUP(value, table, row, exact)", "Recherche horizontale", "=HLOOKUP(2024,A1:E5,3,FALSE)"],
    ["Lookup", "=INDEX(range, row, col)", "Retourne valeur à position", "=INDEX(B:B,5)"],
    ["Lookup", "=MATCH(value, range, type)", "Trouve position", "=MATCH(\"Kinshasa\",B:B,0)"],
    ["Lookup", "=XLOOKUP(value, lookup, return)", "Lookup moderne", "=XLOOKUP(\"Kinshasa\",provinces,Valeur)"],
    ["Logique", "=IF(condition, true, false)", "Condition simple", "=IF(Valeur>1000,\"Élevé\",\"Bas\")"],
    ["Logique", "=IFS(cond1, val1, cond2, val2)", "Conditions multiples", "=IFS(Valeur>1000,\"A\",Valeur>500,\"B\")"],
    ["Logique", "=SWITCH(value, match1, result1...)", "Switch/Case", "=SWITCH(province,\"Kinshasa\",1,\"Ituri\",2)"],
    ["Texte", "=CONCATENATE(text1, text2)", "Concaténation", "=CONCATENATE(province,\"-\",annee)"],
    ["Texte", "=LEFT(text, n)", "N caractères gauche", "=LEFT(province,3)"],
    ["Texte", "=TEXT(value, format)", "Formatage texte", "=TEXT(Valeur,\"#,##0\")"],
    ["Date", "=YEAR(date)", "Extraire année", "=YEAR(TODAY())"],
    ["Date", "=DATEDIF(start, end, unit)", "Différence dates", "=DATEDIF(A1,B1,\"Y\")"],
    ["Avancé", "=UNIQUE(range)", "Valeurs uniques", "=UNIQUE(provinces)"],
    ["Avancé", "=FILTER(range, criteria)", "Filtrage dynamique", "=FILTER(A:H,annees=2024)"],
    ["Avancé", "=SORT(range, col, order)", "Tri dynamique", "=SORT(A:H,8,-1)"],
]

for r_idx, row in enumerate(formulas_list, 3):
    for c_idx, val in enumerate(row, 1):
        cell = ws_formulas.cell(row=r_idx, column=c_idx, value=val)
        cell.border = border
        if r_idx == 3:
            cell.font = header_font
            cell.fill = header_fill

# Column widths
ws_formulas.column_dimensions['A'].width = 15
ws_formulas.column_dimensions['B'].width = 45
ws_formulas.column_dimensions['C'].width = 30
ws_formulas.column_dimensions['D'].width = 45

# ============================================================
# SHEET 9: Gender Analysis
# ============================================================
print("Creating Gender Analysis sheet...")
ws_gender = wb.create_sheet("Analyse_Genre", 8)

ws_gender['A1'] = "ANALYSE PAR GENRE"
ws_gender['A1'].font = Font(bold=True, size=14, color="2E75B6")
ws_gender.merge_cells('A1:E1')

gender_summary = df.groupby('sexes').agg({
    'Valeur': ['sum', 'mean', 'count']
}).round(2)
gender_summary.columns = ['Total', 'Moyenne', 'Nb_Records']
gender_summary = gender_summary.reset_index()
gender_summary['sexes'] = gender_summary['sexes'].fillna('Non spécifié')
gender_summary['Pourcentage'] = (gender_summary['Total'] / gender_summary['Total'].sum() * 100).round(2)

add_data_with_style(ws_gender, gender_summary, start_row=3)

# Add pie chart
ws_gender['A10'] = "Répartition par Genre"
ws_gender['A10'].font = Font(bold=True, size=12)

pie = PieChart()
pie.title = "Distribution par Genre"
data = Reference(ws_gender, min_col=2, min_row=3, max_row=3 + len(gender_summary))
labels = Reference(ws_gender, min_col=1, min_row=4, max_row=3 + len(gender_summary))
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.width = 12
pie.height = 10
ws_gender.add_chart(pie, "A11")

# Column widths
for col in ['A', 'B', 'C', 'D', 'E']:
    ws_gender.column_dimensions[col].width = 18

# ============================================================
# SHEET 10: Age Group Analysis
# ============================================================
print("Creating Age Group Analysis sheet...")
ws_age = wb.create_sheet("Analyse_Age", 9)

ws_age['A1'] = "ANALYSE PAR TRANCHE D'ÂGE"
ws_age['A1'].font = Font(bold=True, size=14, color="2E75B6")
ws_age.merge_cells('A1:E1')

age_summary = df.groupby('tranches_ages').agg({
    'Valeur': ['sum', 'mean', 'count']
}).round(2)
age_summary.columns = ['Total', 'Moyenne', 'Nb_Records']
age_summary = age_summary.reset_index()
age_summary['tranches_ages'] = age_summary['tranches_ages'].fillna('Non spécifié')
age_summary['Pourcentage'] = (age_summary['Total'] / age_summary['Total'].sum() * 100).round(2)

add_data_with_style(ws_age, age_summary, start_row=3)

# Add bar chart
chart = BarChart()
chart.type = "col"
chart.title = "Distribution par Tranche d'Âge"
chart.y_axis.title = "Valeur Totale"

data = Reference(ws_age, min_col=2, min_row=3, max_row=3 + len(age_summary))
cats = Reference(ws_age, min_col=1, min_row=4, max_row=3 + len(age_summary))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width = 15
chart.height = 10
ws_age.add_chart(chart, "G3")

# Column widths
for col in ['A', 'B', 'C', 'D', 'E']:
    ws_age.column_dimensions[col].width = 18
ws_age.column_dimensions['A'].width = 20

# ============================================================
# Save the workbook
# ============================================================
print("\nSaving workbook...")
output_file = 'datavih_analysis.xlsx'
wb.save(output_file)
print(f"✓ Workbook saved as: {output_file}")

print("\n" + "="*60)
print("WORKSHEETS CREATED:")
print("="*60)
sheets = [
    "1. Statistiques_Resume - Summary statistics",
    "2. Resume_Province - Province summary with chart",
    "3. Resume_Annee - Year summary with trend chart",
    "4. Cascade_ONUSIDA - UNAIDS 95-95-95 cascade",
    "5. Tableau_Croise - Pivot table with conditional formatting",
    "6. Reference_Lookup - VLOOKUP/INDEX-MATCH reference tables",
    "7. Dashboard - KPI dashboard with multiple charts",
    "8. Formules_Analyse - Excel formula reference guide",
    "9. Analyse_Genre - Gender analysis with pie chart",
    "10. Analyse_Age - Age group analysis with bar chart",
]
for sheet in sheets:
    print(f"  ✓ {sheet}")

print("\n" + "="*60)
print("EXCEL SKILLS DEMONSTRATED:")
print("="*60)
skills = [
    "- Pivot Tables (manual implementation)",
    "- VLOOKUP/INDEX-MATCH reference tables",
    "- SUMIF/COUNTIF formulas",
    "- Conditional Formatting (Color Scales)",
    "- Data Tables with styling",
    "- Bar Charts, Line Charts, Pie Charts",
    "- KPI Dashboard design",
    "- Data Aggregation & Summarization",
    "- Year-over-Year calculations",
    "- Statistical analysis (mean, median, std)",
]
for skill in skills:
    print(f"  {skill}")

print("\nDone!")
